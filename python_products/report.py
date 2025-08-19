import pandas as pd
import glob
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import logging

# --- 0. ログ設定 ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelName)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# --- 1. 出力ディレクトリの存在チェック ---
save_dir = "data/output"
os.makedirs(save_dir, exist_ok=True)
logging.info(f"出力ディレクトリ: {save_dir}")

# --- 2. データ読み込み ---
files = glob.glob("data/input/*.csv")
if not files:
    raise FileNotFoundError("❌ data/input/ にCSVファイルが見つかりませんでした。")

all_data = []
for file in files:
    basename = os.path.basename(file)
    # ファイル名パターン: sales_tokyo_202501.csv
    match = re.match(r"sales_(.+?)_\d+\.csv", basename)
    branch = match.group(1) if match else "unknown"
    
    df = pd.read_csv(file)
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")  # 不正日付はNaT
    df["支店"] = branch
    all_data.append(df)
    logging.info(f"読込完了: {basename} → 支店={branch}, 行数={len(df)}")

# --- 3. 全データ結合 ---
df_all = pd.concat(all_data, ignore_index=True)

# --- 4. 支店ごとに集計 ---
summary = df_all.groupby("支店")["売上金額"].agg(["sum", "mean"]).reset_index()
summary["mean"] = summary["mean"].round(0).astype(int)  # 四捨五入
summary.columns = ["支店", "売上合計", "売上平均"]

# --- 5. Excel出力 ---
today = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
file_path = os.path.join(save_dir, f"sales_report_{today}.xlsx")

with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    df_all.to_excel(writer, sheet_name="raw_data", index=False)
    summary.to_excel(writer, sheet_name="summary", index=False)

logging.info(f"Excel出力完了: {file_path}")

# --- 6. 条件付き書式（平均以下を赤塗り） ---
wb = load_workbook(file_path)
ws = wb["summary"]

mean_col = 3  # 「売上平均」の列番号
overall_mean = summary["売上平均"].mean()

highlight_threshold = overall_mean
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=mean_col)
    if cell.value < highlight_threshold:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

wb.save(file_path)

logging.info(f"✅ 集計レポートを保存しました → {file_path}")
